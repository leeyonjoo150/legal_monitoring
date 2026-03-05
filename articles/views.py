import json
from datetime import datetime, timedelta

from django.core.paginator import Paginator
from django.db.models import Count
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from articles.models import Article, ArticleMemo, SkippedURL
from articles.progress import get_progress

ARTICLES_PER_PAGE = 30


def _get_filtered_queryset(request):
    """요청 파라미터로 필터링된 QuerySet 반환."""
    qs = Article.objects.all()

    suitability = request.GET.get('suitability')
    if suitability == 'High+Medium':
        qs = qs.filter(suitability__in=['High', 'Medium'])
    elif suitability:
        qs = qs.filter(suitability=suitability)

    case_category = request.GET.get('case_category')
    if case_category:
        qs = qs.filter(case_category=case_category)

    stage = request.GET.get('stage')
    if stage:
        qs = qs.filter(stage=stage)

    title = request.GET.get('title')
    if title:
        qs = qs.filter(title__icontains=title)

    region = request.GET.get('region')
    if region:
        qs = qs.filter(region=region)

    date_from = request.GET.get('date_from')
    if date_from:
        qs = qs.filter(published_at__gte=datetime.fromisoformat(date_from))

    date_to = request.GET.get('date_to')
    if date_to:
        qs = qs.filter(published_at__lte=datetime.fromisoformat(date_to + 'T23:59:59'))

    is_reviewed = request.GET.get('is_reviewed')
    if is_reviewed == 'true':
        qs = qs.filter(is_reviewed=True)
    elif is_reviewed == 'false':
        qs = qs.filter(is_reviewed=False)

    review_passed = request.GET.get('review_passed')
    if review_passed == 'true':
        qs = qs.filter(review_passed=True)
    elif review_passed == 'false':
        qs = qs.filter(review_passed=False)
    elif review_passed == 'null':
        qs = qs.filter(review_passed__isnull=True)

    return qs


def dashboard(request):
    """메인 대시보드 뷰."""
    articles = _get_filtered_queryset(request)

    today = timezone.now().date()
    today_articles = Article.objects.filter(collected_at__date=today)

    categories = (
        Article.objects.values_list('case_category', flat=True)
        .distinct()
        .order_by('case_category')
    )

    paginator = Paginator(articles, ARTICLES_PER_PAGE)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'articles': page_obj,
        'page_obj': page_obj,
        'total_today': today_articles.count(),
        'high_today': today_articles.filter(suitability='High').count(),
        'medium_today': today_articles.filter(suitability='Medium').count(),
        'categories': categories,
        'current_filters': {
            'suitability': request.GET.get('suitability', ''),
            'case_category': request.GET.get('case_category', ''),
            'stage': request.GET.get('stage', ''),
            'region': request.GET.get('region', ''),
            'title': request.GET.get('title', ''),
            'date_from': request.GET.get('date_from', ''),
            'date_to': request.GET.get('date_to', ''),
            'is_reviewed': request.GET.get('is_reviewed', ''),
            'review_passed': request.GET.get('review_passed', ''),
        },
    }
    response = render(request, 'articles/dashboard.html', context)
    response['Cache-Control'] = 'no-store'
    return response


def detail(request, article_id):
    """기사 상세 뷰."""
    article = get_object_or_404(Article, pk=article_id)
    related_articles = article.duplicate_articles.filter(
        related_article__isnull=False
    ).order_by('-skipped_at')
    return render(request, 'articles/detail.html', {
        'article': article,
        'related_articles': related_articles,
    })


def export_xlsx(request):
    """현재 필터 조건으로 엑셀 내보내기."""
    articles = _get_filtered_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = '소송금융 모니터링'

    headers = [
        '제목', '언론사', '게재일', '적합도', '판단 근거',
        '사건 분야', '국내/해외', '상대방', '피해 규모', '진행 단계',
        '진행 단계 상세', '요약', '원문 링크',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for article in articles:
        ws.append([
            article.title,
            article.press,
            article.published_at.strftime('%Y-%m-%d %H:%M'),
            article.suitability,
            article.suitability_reason,
            article.case_category,
            article.region,
            article.defendant,
            article.damage_scale,
            article.stage,
            article.stage_detail,
            article.summary,
            article.url,
        ])

    column_widths = [50, 15, 18, 10, 40, 15, 10, 20, 25, 15, 25, 50, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'legal_radar_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework.decorators import api_view
from rest_framework.response import Response


@swagger_auto_schema(
    method='get',
    operation_summary='분석 진행상황 조회',
    operation_description='현재 분석 진행 상태, 시간 정보, 통계를 반환합니다.',
    responses={200: openapi.Response(
        '진행상황',
        schema=openapi.Schema(type=openapi.TYPE_OBJECT, properties={
            'is_running': openapi.Schema(type=openapi.TYPE_BOOLEAN),
            'phase': openapi.Schema(type=openapi.TYPE_STRING, enum=['collecting', 'analyzing', 'done', '']),
            'current': openapi.Schema(type=openapi.TYPE_INTEGER),
            'total': openapi.Schema(type=openapi.TYPE_INTEGER),
            'started_at': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
            'estimated_end': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
            'current_title': openapi.Schema(type=openapi.TYPE_STRING),
            'saved': openapi.Schema(type=openapi.TYPE_INTEGER),
            'skipped_duplicate': openapi.Schema(type=openapi.TYPE_INTEGER),
            'failed': openapi.Schema(type=openapi.TYPE_INTEGER),
            'last_finished_at': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
            'previous_finished_at': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
            'next_scheduled_at': openapi.Schema(type=openapi.TYPE_STRING, nullable=True),
        }),
    )},
)
@api_view(['GET'])
def progress_status(request):
    """분석 진행상황 JSON API."""
    return Response(get_progress())


@swagger_auto_schema(
    method='post',
    operation_summary='심사 상태 토글',
    operation_description='기사의 심사 완료 여부 또는 통과 여부를 토글합니다.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['action'],
        properties={
            'action': openapi.Schema(type=openapi.TYPE_STRING, enum=['reviewed', 'passed'],
                                     description="'reviewed': 심사 완료 토글, 'passed': 통과 여부 설정"),
            'value': openapi.Schema(type=openapi.TYPE_BOOLEAN, nullable=True,
                                    description="action='passed'일 때 통과 값 (true/false/null)"),
        },
    ),
    responses={200: openapi.Response('결과', schema=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'is_reviewed': openapi.Schema(type=openapi.TYPE_BOOLEAN),
            'review_passed': openapi.Schema(type=openapi.TYPE_BOOLEAN, nullable=True),
        },
    ))},
)
@api_view(['POST'])
def toggle_review(request, article_id):
    """심사 완료 / 통과 여부 토글 API."""
    article = get_object_or_404(Article, pk=article_id)
    data = request.data
    action = data.get('action')

    if action == 'reviewed':
        article.is_reviewed = not article.is_reviewed
        article.save(update_fields=['is_reviewed'])
    elif action == 'passed':
        article.review_passed = data.get('value')
        article.save(update_fields=['review_passed'])

    return Response({
        'is_reviewed': article.is_reviewed,
        'review_passed': article.review_passed,
    })


@swagger_auto_schema(
    method='post',
    operation_summary='적합도 수정',
    operation_description='기사의 투자 적합도를 High / Medium / Low 중 하나로 변경합니다.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['suitability'],
        properties={
            'suitability': openapi.Schema(type=openapi.TYPE_STRING, enum=['High', 'Medium', 'Low']),
        },
    ),
    responses={
        200: openapi.Response('결과', schema=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={'suitability': openapi.Schema(type=openapi.TYPE_STRING)},
        )),
        400: '잘못된 적합도 값',
    },
)
@api_view(['POST'])
def update_suitability(request, article_id):
    """적합도 수정 API."""
    article = get_object_or_404(Article, pk=article_id)
    new_suitability = request.data.get('suitability', '')

    if new_suitability not in ('High', 'Medium', 'Low'):
        return Response({'error': 'invalid suitability'}, status=400)

    article.suitability = new_suitability
    article.save(update_fields=['suitability'])

    return Response({'suitability': article.suitability})


@swagger_auto_schema(
    method='get',
    operation_summary='최신 기사 목록 조회',
    operation_description='after_id 이후 저장된 신규 기사와 오늘의 통계를 반환합니다.',
    manual_parameters=[
        openapi.Parameter('after_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                          description='이 ID 이후의 기사만 반환', default=0),
        openapi.Parameter('suitability', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='적합도 필터 (High/Medium/Low/High+Medium)', required=False),
        openapi.Parameter('case_category', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='사건 분야 필터', required=False),
        openapi.Parameter('stage', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='진행 단계 필터', required=False),
        openapi.Parameter('region', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='국내/해외 필터', required=False),
        openapi.Parameter('title', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='제목 검색 (부분 일치)', required=False),
        openapi.Parameter('date_from', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='게재일 시작 (YYYY-MM-DD)', required=False),
        openapi.Parameter('date_to', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='게재일 종료 (YYYY-MM-DD)', required=False),
        openapi.Parameter('is_reviewed', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='심사 여부 필터 (true/false)', required=False),
        openapi.Parameter('review_passed', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='통과 여부 필터 (true/false/null)', required=False),
    ],
    responses={200: openapi.Response('기사 목록 + 통계', schema=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'articles': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'suitability': openapi.Schema(type=openapi.TYPE_STRING),
                    'region': openapi.Schema(type=openapi.TYPE_STRING),
                    'title': openapi.Schema(type=openapi.TYPE_STRING),
                    'case_category': openapi.Schema(type=openapi.TYPE_STRING),
                    'defendant': openapi.Schema(type=openapi.TYPE_STRING),
                    'stage': openapi.Schema(type=openapi.TYPE_STRING),
                    'press': openapi.Schema(type=openapi.TYPE_STRING),
                    'published_at': openapi.Schema(type=openapi.TYPE_STRING),
                },
            )),
            'total_today': openapi.Schema(type=openapi.TYPE_INTEGER),
            'high_today': openapi.Schema(type=openapi.TYPE_INTEGER),
            'medium_today': openapi.Schema(type=openapi.TYPE_INTEGER),
        },
    ))},
)
@api_view(['GET'])
def api_articles_latest(request):
    """after_id 이후 저장된 신규 기사 + 오늘 통계 반환."""
    try:
        after_id = int(request.GET.get('after_id', 0))
    except (ValueError, TypeError):
        after_id = 0

    new_articles = _get_filtered_queryset(request).filter(id__gt=after_id).order_by('-id')[:50]

    today = timezone.now().date()
    today_qs = Article.objects.filter(collected_at__date=today)

    return Response({
        'articles': [
            {
                'id': a.id,
                'suitability': a.suitability,
                'region': a.region,
                'title': a.title,
                'case_category': a.case_category,
                'defendant': a.defendant,
                'stage': a.stage,
                'press': a.press,
                'published_at': a.published_at.strftime('%Y-%m-%d %H:%M'),
            }
            for a in new_articles
        ],
        'total_today': today_qs.count(),
        'high_today': today_qs.filter(suitability='High').count(),
        'medium_today': today_qs.filter(suitability='Medium').count(),
    })


@swagger_auto_schema(
    method='get',
    operation_summary='차트 데이터 조회',
    operation_description='진행 단계별 High 기사 수, 심사 현황, 중복 Top 10을 반환합니다.',
    manual_parameters=[
        openapi.Parameter('days', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                          description='조회 기간 (일 수, 기본값 30, 최대 365)', required=False, default=30),
    ],
    responses={200: openapi.Response('차트 데이터', schema=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'stage_high': openapi.Schema(type=openapi.TYPE_OBJECT, properties={
                'labels': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_STRING)),
                'counts': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            }),
            'review_status': openapi.Schema(type=openapi.TYPE_OBJECT, properties={
                'reviewed': openapi.Schema(type=openapi.TYPE_INTEGER),
                'unreviewed': openapi.Schema(type=openapi.TYPE_INTEGER),
            }),
            'top10_duplicates': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(
                type=openapi.TYPE_OBJECT,
                properties={
                    'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                    'title': openapi.Schema(type=openapi.TYPE_STRING),
                    'count': openapi.Schema(type=openapi.TYPE_INTEGER),
                },
            )),
        },
    ))},
)
@api_view(['GET'])
def api_charts(request):
    """차트 데이터 API."""
    try:
        days = int(request.GET.get('days', 30))
        if days < 1:
            days = 1
        elif days > 365:
            days = 365
    except (ValueError, TypeError):
        days = 30

    since = timezone.now() - timedelta(days=days)

    # 진행 단계별 High 기사 수
    STAGES = ['피해 발생', '관련 절차 진행', '소송중', '판결 선고', '종결']
    high_qs = Article.objects.filter(suitability='High', published_at__gte=since)
    stage_counts = [high_qs.filter(stage=s).count() for s in STAGES]

    # 심사 현황 (High + Medium)
    review_qs = Article.objects.filter(
        suitability__in=['High', 'Medium'],
        published_at__gte=since,
    )
    reviewed = review_qs.filter(is_reviewed=True).count()
    unreviewed = review_qs.filter(is_reviewed=False).count()

    # 중복 수 Top 10 (High only)
    top10 = (
        SkippedURL.objects
        .filter(
            related_article__isnull=False,
            related_article__suitability='High',
            related_article__published_at__gte=since,
        )
        .values('related_article_id', 'related_article__title')
        .annotate(cnt=Count('id'))
        .order_by('-cnt')[:10]
    )

    return Response({
        'stage_high': {
            'labels': STAGES,
            'counts': stage_counts,
        },
        'review_status': {
            'reviewed': reviewed,
            'unreviewed': unreviewed,
        },
        'top10_duplicates': [
            {
                'id': item['related_article_id'],
                'title': item['related_article__title'],
                'count': item['cnt'],
            }
            for item in top10
        ],
    })


@swagger_auto_schema(
    method='post',
    operation_summary='메모 추가',
    operation_description='기사에 메모를 추가합니다.',
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        required=['author', 'content'],
        properties={
            'author': openapi.Schema(type=openapi.TYPE_STRING, description='작성자'),
            'content': openapi.Schema(type=openapi.TYPE_STRING, description='메모 내용'),
        },
    ),
    responses={
        200: openapi.Response('추가된 메모', schema=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                'author': openapi.Schema(type=openapi.TYPE_STRING),
                'content': openapi.Schema(type=openapi.TYPE_STRING),
                'created_at': openapi.Schema(type=openapi.TYPE_STRING),
            },
        )),
        400: '작성자 또는 내용 누락',
    },
)
@api_view(['POST'])
def add_memo(request, article_id):
    """메모 추가 API."""
    article = get_object_or_404(Article, pk=article_id)
    author = request.data.get('author', '').strip()
    content = request.data.get('content', '').strip()

    if not content:
        return Response({'error': '내용을 입력해주세요.'}, status=400)
    if not author:
        return Response({'error': '작성자를 입력해주세요.'}, status=400)

    memo = ArticleMemo.objects.create(
        article=article,
        author=author,
        content=content,
    )

    return Response({
        'id': memo.id,
        'author': memo.author,
        'content': memo.content,
        'created_at': memo.created_at.strftime('%Y-%m-%d %H:%M'),
    })


@swagger_auto_schema(
    method='get',
    operation_summary='PDF용 기사 목록 조회',
    operation_description='PDF 내보내기 팝업용 기사 목록을 반환합니다. 현재 필터 조건과 max_id 스냅샷이 적용됩니다.',
    manual_parameters=[
        openapi.Parameter('max_id', openapi.IN_QUERY, type=openapi.TYPE_INTEGER,
                          description='이 ID 이하의 기사만 반환 (스냅샷 기준)', required=False),
        openapi.Parameter('suitability', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('case_category', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('stage', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('region', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('title', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('date_from', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('date_to', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('is_reviewed', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
        openapi.Parameter('review_passed', openapi.IN_QUERY, type=openapi.TYPE_STRING, required=False),
    ],
    responses={200: openapi.Response('기사 목록', schema=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'articles': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
            'total': openapi.Schema(type=openapi.TYPE_INTEGER),
        },
    ))},
)
@api_view(['GET'])
def api_pdf_articles(request):
    """PDF 팝업용 기사 목록 (필터 + max_id 스냅샷 적용)."""
    try:
        max_id = int(request.GET.get('max_id', 0))
    except (ValueError, TypeError):
        max_id = 0

    qs = _get_filtered_queryset(request)
    if max_id > 0:
        qs = qs.filter(id__lte=max_id)

    articles = list(qs.values(
        'id', 'suitability', 'region', 'title', 'case_category',
        'defendant', 'stage', 'press', 'published_at',
    ))

    for a in articles:
        if hasattr(a['published_at'], 'strftime'):
            a['published_at'] = a['published_at'].strftime('%Y-%m-%d %H:%M')

    return Response({
        'articles': articles,
        'total': len(articles),
    })


@swagger_auto_schema(
    method='get',
    operation_summary='PDF용 기사 상세 조회',
    operation_description='PDF 내보내기용 선택 기사의 전체 필드와 메모를 반환합니다.',
    manual_parameters=[
        openapi.Parameter('ids', openapi.IN_QUERY, type=openapi.TYPE_STRING,
                          description='쉼표로 구분된 기사 ID 목록 (예: 1,2,3)', required=True),
    ],
    responses={
        200: openapi.Response('기사 상세 목록', schema=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                'articles': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'id': openapi.Schema(type=openapi.TYPE_INTEGER),
                        'title': openapi.Schema(type=openapi.TYPE_STRING),
                        'press': openapi.Schema(type=openapi.TYPE_STRING),
                        'published_at': openapi.Schema(type=openapi.TYPE_STRING),
                        'ai_suitability': openapi.Schema(type=openapi.TYPE_STRING),
                        'suitability': openapi.Schema(type=openapi.TYPE_STRING),
                        'suitability_reason': openapi.Schema(type=openapi.TYPE_STRING),
                        'case_category': openapi.Schema(type=openapi.TYPE_STRING),
                        'defendant': openapi.Schema(type=openapi.TYPE_STRING),
                        'damage_scale': openapi.Schema(type=openapi.TYPE_STRING),
                        'stage': openapi.Schema(type=openapi.TYPE_STRING),
                        'stage_detail': openapi.Schema(type=openapi.TYPE_STRING),
                        'summary': openapi.Schema(type=openapi.TYPE_STRING),
                        'region': openapi.Schema(type=openapi.TYPE_STRING),
                        'memos': openapi.Schema(type=openapi.TYPE_ARRAY, items=openapi.Schema(type=openapi.TYPE_OBJECT)),
                    },
                )),
            },
        )),
        400: 'ids 파라미터 누락 또는 형식 오류',
    },
)
@api_view(['GET'])
def api_pdf_article_details(request):
    """PDF용 선택 기사 상세 데이터 (전체 필드 + 메모)."""
    ids_str = request.GET.get('ids', '')
    if not ids_str:
        return Response({'articles': []})

    try:
        ids = [int(i) for i in ids_str.split(',') if i.strip()]
    except ValueError:
        return Response({'error': 'invalid ids'}, status=400)

    articles_qs = Article.objects.filter(id__in=ids).prefetch_related('memos')

    id_to_article = {}
    for article in articles_qs:
        id_to_article[article.id] = {
            'id': article.id,
            'title': article.title,
            'press': article.press,
            'published_at': article.published_at.strftime('%Y-%m-%d %H:%M'),
            'ai_suitability': article.ai_suitability,
            'suitability': article.suitability,
            'suitability_reason': article.suitability_reason,
            'case_category': article.case_category,
            'defendant': article.defendant,
            'damage_scale': article.damage_scale,
            'stage': article.stage,
            'stage_detail': article.stage_detail,
            'summary': article.summary,
            'region': article.region,
            'memos': [
                {
                    'author': m.author,
                    'content': m.content,
                    'created_at': m.created_at.strftime('%Y-%m-%d %H:%M'),
                }
                for m in article.memos.all()
            ],
        }

    # 선택 순서 유지
    ordered = [id_to_article[i] for i in ids if i in id_to_article]

    return Response({'articles': ordered})
